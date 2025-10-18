"""Bootstrap or text-mine fail-code dictionary entries.

Supports two modes:
  * table (default): read structured columns with explicit codes/names.
  * text-extract: mine candidate codes/names from free-text columns.

Designed to operate offline inside the Windows bundle produced by this repo.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections import defaultdict
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path

import click
import yaml
from openpyxl import load_workbook

from app import create_app, db
from app.models import FailCode

# -----------------------------------------------------------------------------
# Defaults & configuration helpers
# -----------------------------------------------------------------------------

DEFAULT_CODE_PATTERNS = [r"\b[0-9]{3,6}\b", r"\b[A-Z]{2,6}\b"]
DEFAULT_DENY_PATTERNS = [
    r"^M\d{3}$",
    r"^R\d{3}$",
    r"^SN\d+$",
    r"^X0[A-Z0-9]+$",
    r"^\d+F/\d+ea$",
    r"^\d+(\.\d+)?h$",
]
DEFAULT_STOPWORDS = {
    "DOE",
    "PFQ",
    "PRQ",
    "AQL",
    "TC",
    "SREP",
    "HOLD",
    "AGING",
    "REPAIR",
    "QA",
    "PROCESS",
    "PASS",
    "RETEST",
    "ALL",
    "CLOSE",
}
CHINESE_FAILURE_TERMS = [
    "不识别",
    "功能性不良",
    "开路",
    "短路",
    "坏块",
    "ECC错误",
    "掉电",
    "过温",
    "老化",
    "裂纹",
    "腐蚀",
    "焊点",
    "掉颗",
    "CTRL功能性不良",
]
DEFAULT_CONTEXT_CHARS = 30
DEFAULT_ENCODING = "utf-8"

TEXT_MODE = "text-extract"
TABLE_MODE = "table"


@dataclass
class FileStats:
    path: Path
    rows_total: int = 0
    rows_with_code: int = 0
    tokens_found: int = 0
    missing_code_rows: list[dict[str, object]] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return {
            "file": str(self.path),
            "rows_total": self.rows_total,
            "rows_with_code": self.rows_with_code,
            "tokens_found": self.tokens_found,
            "missing_code_rows": len(self.missing_code_rows),
        }


@dataclass
class FailCodeAggregate:
    code: str
    name_candidates: set[str] = field(default_factory=set)
    description_candidates: set[str] = field(default_factory=set)
    sources: set[str] = field(default_factory=set)
    missing_name_count: int = 0


@dataclass
class RawToken:
    source_file: str
    row_id: int
    column: str
    token: str
    token_type: str  # "code" or "name"
    canonical_code: str
    context: str


@dataclass
class TokenAggregate:
    token: str
    token_type: str
    canonical_code: str
    occurrences: int = 0
    sources: set[str] = field(default_factory=set)
    rows: set[int] = field(default_factory=set)


# -----------------------------------------------------------------------------
# File parsing utilities
# -----------------------------------------------------------------------------


def iter_csv_rows(
    path: Path,
    delimiter: str | None = None,
    encoding: str = DEFAULT_ENCODING,
) -> Iterator[tuple[int, dict[str, object]]]:
    with path.open("r", newline="", encoding=encoding) as handle:
        if delimiter is None:
            sniffer = csv.Sniffer()
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = sniffer.sniff(sample)
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
        else:
            reader = csv.DictReader(handle, delimiter=delimiter)
        yield from enumerate(reader, start=2)


def iter_xlsx_rows(path: Path) -> Iterator[tuple[int, dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for idx, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            values = [
                row[col] if col < len(row) else None for col in range(len(headers))
            ]
            yield idx, {headers[i]: values[i] for i in range(len(headers))}
    finally:
        workbook.close()


def iter_rows(
    path: Path,
    delimiter: str | None = None,
    encoding: str = DEFAULT_ENCODING,
) -> Iterator[tuple[int, dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_rows(path, delimiter, encoding)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from iter_xlsx_rows(path)
    else:
        raise click.BadParameter(f"Unsupported file type: {path.suffix}")


# -----------------------------------------------------------------------------
# General helpers
# -----------------------------------------------------------------------------


def normalize_text(value: object | None, upper: bool = False) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.upper() if upper else text


def load_stopwords(file_path: Path | None) -> set[str]:
    words = set(DEFAULT_STOPWORDS)
    if file_path is None:
        return words
    with file_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            token = line.strip()
            if token and not token.startswith("#"):
                words.add(token.upper())
    return words


def load_config(config_path: Path | None) -> dict[str, object]:
    config: dict[str, object] = {}
    if config_path and config_path.exists():
        with config_path.open("r", encoding="utf-8") as handle:
            loaded = yaml.safe_load(handle) or {}
            if not isinstance(loaded, dict):
                raise click.BadParameter("Configuration file must define a mapping")
            config = loaded
    return config


def compile_patterns(patterns: Iterable[str]) -> list[re.Pattern[str]]:
    compiled: list[re.Pattern[str]] = []
    for pattern in patterns:
        try:
            compiled.append(re.compile(pattern))
        except re.error as exc:
            raise click.BadParameter(
                f"Invalid regex pattern '{pattern}': {exc}"
            ) from exc
    return compiled


def build_name_pattern(extra_terms: Iterable[str]) -> re.Pattern[str]:
    phrases = list(CHINESE_FAILURE_TERMS)
    phrases.extend(term for term in extra_terms if isinstance(term, str) and term)
    escaped = [re.escape(term) for term in phrases]
    combined = "|".join(sorted(set(escaped), key=len, reverse=True))
    return re.compile(combined)


def slugify_legacy(token: str) -> str:
    upper = token.upper()
    ascii_slug = re.sub(r"[^A-Z0-9]+", "-", upper).strip("-")
    if ascii_slug:
        slug = ascii_slug[:48]
    else:
        digest = hashlib.sha1(token.encode("utf-8")).hexdigest().upper()[:10]
        slug = f"ZH{digest}"
    return f"LEGACY-{slug}"


def ensure_fail_code(
    code: str,
    short_name: str | None,
    is_provisional: bool,
    source: str | None,
) -> tuple[bool, bool]:
    created = False
    updated = False
    record = FailCode.query.filter_by(code=code).one_or_none()
    if record is None:
        record = FailCode(
            code=code,
            short_name=short_name,
            is_provisional=is_provisional,
            source=source,
        )
        db.session.add(record)
        created = True
    else:
        if short_name and not record.short_name:
            record.short_name = short_name
            updated = True
        if source and not record.source:
            record.source = source
            updated = True
        if is_provisional and not record.is_provisional:
            record.is_provisional = True
            updated = True
    return created, updated


# -----------------------------------------------------------------------------
# Structured table mode
# -----------------------------------------------------------------------------


def run_table_mode(
    inputs: tuple[Path, ...],
    code_column: str,
    name_column: str | None,
    description_column: str | None,
    delimiter: str | None,
    encoding: str,
    output_dir: Path,
    descriptions_preferred: bool,
) -> None:
    aggregates: dict[str, FailCodeAggregate] = {}
    stats: list[FileStats] = []
    conflicts: dict[str, set[str]] = defaultdict(set)
    extraction_errors: list[dict[str, object]] = []

    with AppContext():
        for path in inputs:
            file_stats = FileStats(path=path)
            stats.append(file_stats)
            try:
                for row_index, row in iter_rows(path, delimiter, encoding):
                    file_stats.rows_total += 1
                    raw_code = row.get(code_column)
                    normalized_code = normalize_text(raw_code, upper=True)
                    if not normalized_code:
                        file_stats.missing_code_rows.append(
                            {
                                "file": str(path),
                                "row": row_index,
                                "reason": "missing fail code",
                            }
                        )
                        continue

                    file_stats.rows_with_code += 1
                    aggregate = aggregates.setdefault(
                        normalized_code,
                        FailCodeAggregate(code=normalized_code),
                    )
                    aggregate.sources.add(str(path))

                    if name_column:
                        raw_name = row.get(name_column)
                        normalized_name = normalize_text(raw_name)
                        if normalized_name:
                            aggregate.name_candidates.add(normalized_name)
                        else:
                            aggregate.missing_name_count += 1
                    else:
                        aggregate.missing_name_count += 1

                    if description_column:
                        raw_description = row.get(description_column)
                        normalized_description = normalize_text(raw_description)
                        if normalized_description:
                            aggregate.description_candidates.add(normalized_description)

                    if name_column and len(aggregate.name_candidates) > 1:
                        conflicts[normalized_code] = set(aggregate.name_candidates)

                extraction_errors.extend(file_stats.missing_code_rows)
            except Exception as exc:  # noqa: BLE001
                extraction_errors.append(
                    {
                        "file": str(path),
                        "row": None,
                        "reason": f"failed to parse file: {exc}",
                    }
                )
                click.echo(f"[ERROR] {path}: {exc}")

        created, updated = upsert_from_aggregates(
            aggregates, descriptions_preferred=descriptions_preferred
        )

    write_structured_reports(
        output_dir,
        aggregates,
        conflicts,
        extraction_errors,
        stats,
        created,
        updated,
        code_column,
        name_column,
        description_column,
    )


def upsert_from_aggregates(
    aggregates: dict[str, FailCodeAggregate],
    descriptions_preferred: bool,
) -> tuple[int, int]:
    codes = list(aggregates.keys())
    existing = {
        entry.code: entry
        for entry in FailCode.query.filter(FailCode.code.in_(codes)).all()
    }

    created = 0
    updated_codes: set[str] = set()

    for code, aggregate in aggregates.items():
        record = existing.get(code)
        primary_name = next((name for name in aggregate.name_candidates if name), None)
        primary_description = None
        if descriptions_preferred:
            primary_description = next(
                (
                    description
                    for description in aggregate.description_candidates
                    if description
                ),
                None,
            )

        if record is None:
            record = FailCode(
                code=code,
                short_name=primary_name,
                description=primary_description,
            )
            db.session.add(record)
            created += 1
        else:
            if primary_name and not record.short_name:
                record.short_name = primary_name
                updated_codes.add(code)
            if primary_description and not record.description:
                record.description = primary_description
                updated_codes.add(code)

    db.session.commit()
    return created, len(updated_codes)


def write_structured_reports(
    output_dir: Path,
    aggregates: dict[str, FailCodeAggregate],
    conflicts: dict[str, set[str]],
    extraction_errors: list[dict[str, object]],
    stats: list[FileStats],
    created: int,
    updated: int,
    code_column: str,
    name_column: str | None,
    description_column: str | None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    missing_names_path = output_dir / "missing_names.csv"
    with missing_names_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["code", "missing_name_count", "sources"])
        for code, aggregate in sorted(aggregates.items()):
            if aggregate.missing_name_count > 0:
                writer.writerow(
                    [
                        code,
                        aggregate.missing_name_count,
                        "|".join(sorted(aggregate.sources)),
                    ]
                )

    conflicts_path = output_dir / "conflicts.csv"
    with conflicts_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["code", "names", "sources"])
        for code, names in sorted(conflicts.items()):
            writer.writerow(
                [
                    code,
                    "|".join(sorted(names)),
                    "|".join(sorted(aggregates[code].sources)),
                ]
            )

    errors_path = output_dir / "errors.csv"
    with errors_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["file", "row", "reason"])
        for error in extraction_errors:
            writer.writerow([error.get("file"), error.get("row"), error.get("reason")])

    ingestion_log_path = output_dir / "ingestion_log.json"
    with ingestion_log_path.open("w", encoding="utf-8") as handle:
        timestamp = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
        json.dump(
            {
                "generated_at": timestamp,
                "mode": TABLE_MODE,
                "code_column": code_column,
                "name_column": name_column,
                "description_column": description_column,
                "files": [stat.to_dict() for stat in stats],
                "totals": {
                    "unique_codes": len(aggregates),
                    "created": created,
                    "updated": updated,
                    "missing_code_rows": len(extraction_errors),
                },
            },
            handle,
            indent=2,
        )


# -----------------------------------------------------------------------------
# Text extraction mode
# -----------------------------------------------------------------------------


def run_text_mode(
    inputs: tuple[Path, ...],
    text_columns: tuple[str, ...],
    config: dict[str, object],
    stopwords_file: Path | None,
    deny_regex_options: tuple[str, ...],
    delimiter: str | None,
    encoding: str,
    output_dir: Path,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    code_patterns = (
        config.get("code_regexes", DEFAULT_CODE_PATTERNS) or DEFAULT_CODE_PATTERNS
    )
    deny_patterns = list(DEFAULT_DENY_PATTERNS)
    deny_patterns.extend(
        pattern for pattern in config.get("deny_regexes", []) if pattern
    )
    deny_patterns.extend(deny_regex_options)
    stopwords = load_stopwords(stopwords_file)
    stopwords.update({word.upper() for word in config.get("stopwords", [])})
    context_chars = int(config.get("context_chars", DEFAULT_CONTEXT_CHARS))
    extra_terms = config.get("additional_failure_terms", [])
    name_pattern = build_name_pattern(extra_terms)
    compiled_code_patterns = compile_patterns(code_patterns)
    compiled_deny_patterns = compile_patterns(deny_patterns) if deny_patterns else []

    raw_tokens: list[RawToken] = []
    aggregates: dict[str, TokenAggregate] = {}
    alias_map: dict[str, str] = {}
    alias_upper_map: dict[str, str] = {}
    code_to_names: dict[str, set[str]] = defaultdict(set)
    alias_conflicts: set[tuple[str, str, str]] = set()
    stats: list[FileStats] = []

    created_total = 0
    updated_total = 0

    with AppContext():
        for path in inputs:
            file_stats = FileStats(path=path)
            stats.append(file_stats)
            try:
                for row_index, row in iter_rows(path, delimiter, encoding):
                    file_stats.rows_total += 1
                    column_tokens: list[dict[str, object]] = []
                    for column in text_columns:
                        value = row.get(column)
                        if not value:
                            continue
                        text = str(value)
                        column_tokens.extend(
                            extract_tokens_from_text(
                                column,
                                row_index,
                                text,
                                compiled_code_patterns,
                                name_pattern,
                                compiled_deny_patterns,
                                stopwords,
                                context_chars,
                            )
                        )

                    if not column_tokens:
                        continue

                    file_stats.tokens_found += len(column_tokens)
                    for token_info in column_tokens:
                        token = token_info["token"]
                        token_type = token_info["token_type"]
                        context = token_info["context"]
                        column = token_info["column"]

                        if token_type == "code":
                            canonical_code = token.upper()
                            code_to_names.setdefault(canonical_code, set())
                            created, updated = ensure_fail_code(
                                canonical_code,
                                short_name=None,
                                is_provisional=True,
                                source="text-extract",
                            )
                            alias_map.setdefault(token, canonical_code)
                            alias_upper_map.setdefault(token.upper(), canonical_code)
                        else:
                            canonical_code = alias_map.get(token)
                            if canonical_code is None:
                                canonical_code = slugify_legacy(token)
                                alias_map[token] = canonical_code
                                alias_upper_map[token.upper()] = canonical_code
                                created, updated = ensure_fail_code(
                                    canonical_code,
                                    short_name=token,
                                    is_provisional=True,
                                    source="legacy-text",
                                )
                            else:
                                created, updated = ensure_fail_code(
                                    canonical_code,
                                    short_name=token,
                                    is_provisional=True,
                                    source="legacy-text",
                                )
                            code_to_names[canonical_code].add(token)

                        created_total += int(created)
                        updated_total += int(updated)

                        alias_key = token.upper()
                        existing_alias = alias_upper_map.get(alias_key)
                        if existing_alias and existing_alias != canonical_code:
                            alias_conflicts.add((token, existing_alias, canonical_code))
                        else:
                            alias_upper_map[alias_key] = canonical_code

                        aggregates_key = (token, token_type)
                        if aggregates_key not in aggregates:
                            aggregates[aggregates_key] = TokenAggregate(
                                token=token,
                                token_type=token_type,
                                canonical_code=canonical_code,
                            )
                        aggreg = aggregates[aggregates_key]
                        aggreg.occurrences += 1
                        aggreg.sources.add(str(path))
                        aggreg.rows.add(row_index)

                        raw_tokens.append(
                            RawToken(
                                source_file=str(path),
                                row_id=row_index,
                                column=column,
                                token=token,
                                token_type=token_type,
                                canonical_code=canonical_code,
                                context=context,
                            )
                        )
            except Exception as exc:  # noqa: BLE001
                click.echo(f"[ERROR] {path}: {exc}")

        db.session.commit()

    write_text_reports(
        output_dir,
        raw_tokens,
        aggregates,
        alias_map,
        code_to_names,
        sorted(alias_conflicts),
        stats,
        created_total,
        updated_total,
        text_columns,
        code_patterns,
        deny_patterns,
        stopwords,
        context_chars,
    )


def extract_tokens_from_text(
    column: str,
    row_index: int,
    full_text: str,
    code_patterns: list[re.Pattern[str]],
    name_pattern: re.Pattern[str],
    deny_patterns: list[re.Pattern[str]],
    stopwords: set[str],
    context_chars: int,
) -> list[dict[str, object]]:
    tokens: list[dict[str, object]] = []
    matched_spans: set[tuple[int, int, str, str]] = set()

    def add_token(start: int, end: int, token: str, token_type: str) -> None:
        span_key = (start, end, token, token_type)
        if span_key in matched_spans:
            return
        matched_spans.add(span_key)
        context_start = max(0, start - context_chars)
        context_end = min(len(full_text), end + context_chars)
        tokens.append(
            {
                "row": row_index,
                "column": column,
                "token": token,
                "token_type": token_type,
                "context": full_text[context_start:context_end],
            }
        )

    for pattern in code_patterns:
        for match in pattern.finditer(full_text):
            token = match.group(0)
            upper = token.upper()
            if upper in stopwords:
                continue
            if any(deny.search(upper) for deny in deny_patterns):
                continue
            add_token(match.start(), match.end(), token, "code")

    for match in name_pattern.finditer(full_text):
        token = match.group(0)
        add_token(match.start(), match.end(), token, "name")

    return tokens


def write_text_reports(
    output_dir: Path,
    raw_tokens: list[RawToken],
    aggregates: dict[tuple[str, str], TokenAggregate],
    alias_map: dict[str, str],
    code_to_names: dict[str, set[str]],
    alias_conflicts: list[tuple[str, str, str]],
    stats: list[FileStats],
    created_total: int,
    updated_total: int,
    text_columns: tuple[str, ...],
    code_patterns: Iterable[str],
    deny_patterns: Iterable[str],
    stopwords: Iterable[str],
    context_chars: int,
) -> None:
    def write_csv(path: Path, rows: Iterable[Iterable[object]]) -> None:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in rows:
                writer.writerow(row)

    raw_path = output_dir / "raw_failure_tokens.csv"
    write_csv(
        raw_path,
        [
            (
                "source_file",
                "row_id",
                "column",
                "token",
                "token_type",
                "canonical_code",
                "context",
            )
        ]
        + [
            (
                token.source_file,
                token.row_id,
                token.column,
                token.token,
                token.token_type,
                token.canonical_code,
                token.context,
            )
            for token in raw_tokens
        ],
    )

    distinct_path = output_dir / "distinct_tokens.csv"
    write_csv(
        distinct_path,
        [("token", "token_type", "canonical_code", "occurrences", "sources")]
        + [
            (
                agg.token,
                agg.token_type,
                agg.canonical_code,
                agg.occurrences,
                "|".join(sorted(agg.sources)),
            )
            for agg in sorted(
                aggregates.values(), key=lambda item: (item.token_type, item.token)
            )
        ],
    )

    aliases_path = output_dir / "aliases.csv"
    write_csv(
        aliases_path,
        [("token", "token_type", "canonical_code", "source")]
        + [
            (
                token,
                "code" if canonical == token.upper() or canonical == token else "name",
                canonical,
                "text-extract"
                if canonical == token.upper() or canonical == token
                else "legacy-text",
            )
            for token, canonical in sorted(alias_map.items())
        ],
    )

    missing_names_path = output_dir / "missing_names.csv"
    write_csv(
        missing_names_path,
        [("code", "occurrences", "sources")]
        + [
            (
                code,
                sum(
                    agg.occurrences
                    for (token, token_type), agg in aggregates.items()
                    if token_type == "code" and token.upper() == code
                ),
                "|".join(
                    sorted(
                        {
                            source
                            for (token, token_type), agg in aggregates.items()
                            if token_type == "code" and token.upper() == code
                            for source in agg.sources
                        }
                    )
                ),
            )
            for code, names in sorted(code_to_names.items())
            if not names
        ],
    )

    conflicts_path = output_dir / "conflicts.csv"
    write_csv(
        conflicts_path,
        [("token", "existing_code", "new_code")] + alias_conflicts,
    )

    ingestion_log_path = output_dir / "ingestion_log.json"
    timestamp = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    with ingestion_log_path.open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "generated_at": timestamp,
                "mode": TEXT_MODE,
                "text_columns": list(text_columns),
                "code_patterns": list(code_patterns),
                "deny_patterns": list(deny_patterns),
                "stopwords_count": len(set(stopwords)),
                "context_chars": context_chars,
                "totals": {
                    "raw_tokens": len(raw_tokens),
                    "distinct_tokens": len(aggregates),
                    "fail_codes_created": created_total,
                    "fail_codes_updated": updated_total,
                },
                "files": [stat.to_dict() for stat in stats],
            },
            handle,
            indent=2,
        )


# -----------------------------------------------------------------------------
# Application context helper
# -----------------------------------------------------------------------------


class AppContext:
    def __init__(self) -> None:
        self.app = create_app(os.getenv("FLASK_ENV", "development"))

    def __enter__(self):
        self._ctx = self.app.app_context()
        self._ctx.push()
        return self._ctx

    def __exit__(self, exc_type, exc, tb) -> None:
        self._ctx.pop()


# -----------------------------------------------------------------------------
# CLI entrypoint
# -----------------------------------------------------------------------------


@click.command()
@click.argument(
    "inputs", nargs=-1, type=click.Path(exists=True, dir_okay=False, path_type=Path)
)
@click.option("--mode", type=click.Choice([TABLE_MODE, TEXT_MODE]), default=TABLE_MODE)
@click.option(
    "--code-column",
    default="fail_code",
    show_default=True,
    help="Structured mode: column containing the fail code",
)
@click.option(
    "--name-column",
    default=None,
    help="Structured mode: optional column containing the fail name",
)
@click.option(
    "--description-column",
    default=None,
    help="Structured mode: optional column containing the fail description",
)
@click.option("--text-col", "text_columns", multiple=True, default=("评价过程",))
@click.option(
    "--delimiter", default=None, help="Optional delimiter override for CSV files"
)
@click.option("--encoding", default=DEFAULT_ENCODING, show_default=True)
@click.option("--config", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--stopwords-file",
    type=click.Path(exists=True, path_type=Path),
    help="Optional custom stopwords file",
)
@click.option("--deny-regex", "deny_regexes", multiple=True)
@click.option(
    "--out-dir",
    "out_dir",
    default=Path("reports/fail_code_bootstrap"),
    type=click.Path(file_okay=False, path_type=Path),
    show_default=True,
    help="Directory where reports are written",
)
@click.option(
    "--output-dir",
    "legacy_output_dir",
    default=None,
    type=click.Path(file_okay=False, path_type=Path),
    help="Deprecated alias for --out-dir",
)
@click.option("--use-description", is_flag=True, default=False)
def main(
    inputs: tuple[Path, ...],
    mode: str,
    code_column: str,
    name_column: str | None,
    description_column: str | None,
    text_columns: tuple[str, ...],
    delimiter: str | None,
    encoding: str,
    config: Path | None,
    stopwords_file: Path | None,
    deny_regexes: tuple[str, ...],
    out_dir: Path,
    legacy_output_dir: Path | None,
    use_description: bool,
) -> None:
    if not inputs:
        raise click.BadParameter("Provide at least one input file")

    output_dir = legacy_output_dir or out_dir
    config_data = load_config(config)

    if mode == TABLE_MODE:
        run_table_mode(
            inputs,
            code_column,
            name_column,
            description_column,
            delimiter,
            encoding,
            output_dir,
            use_description,
        )
    else:
        run_text_mode(
            inputs,
            text_columns,
            config_data,
            stopwords_file,
            deny_regexes,
            delimiter,
            encoding,
            output_dir,
        )


if __name__ == "__main__":
    main()
